"""Generic .xlsx row parser shared by every module's bulk-upload endpoint.

Reads the header row, maps each header to a snake_case key, and yields one
dict per data row tagged with its 1-based Excel row number under the
"_row_number" key (used in bulk-import error messages so an admin can find
the exact row to fix in their sheet). Blank rows are skipped.
"""

import re
from io import BytesIO

from fastapi import HTTPException, UploadFile, status
from openpyxl import load_workbook


def _normalize_header(value: object) -> str:
    return re.sub(r"\s+", "_", str(value).strip().lower())


async def parse_excel_rows(file: UploadFile, required_columns: set[str]) -> list[dict[str, object]]:
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(status.HTTP_422_UNPROCESSABLE_ENTITY, "Only .xlsx files are supported")

    contents = await file.read()
    try:
        workbook = load_workbook(filename=BytesIO(contents), read_only=True, data_only=True)
    except Exception as exc:
        raise HTTPException(status.HTTP_422_UNPROCESSABLE_ENTITY, "Could not read the Excel file") from exc

    sheet = workbook.active
    rows_iter = sheet.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        raise HTTPException(status.HTTP_422_UNPROCESSABLE_ENTITY, "The sheet is empty")

    headers = [_normalize_header(h) for h in header_row if h is not None]
    missing = required_columns - set(headers)
    if missing:
        raise HTTPException(
            status.HTTP_422_UNPROCESSABLE_ENTITY,
            f"Missing required column(s): {', '.join(sorted(missing))}",
        )

    rows: list[dict[str, object]] = []
    for excel_row_num, values in enumerate(rows_iter, start=2):
        if values is None or all(v is None for v in values):
            continue
        # Blank cells are omitted entirely (not passed through as None) so
        # each field's own Pydantic default applies — a bare `None` would
        # otherwise fail validation on non-Optional fields that have a
        # default, e.g. status: str = "pending" or is_new_voter: bool = False.
        row = {
            headers[i]: values[i]
            for i in range(len(headers))
            if i < len(values) and values[i] is not None
        }
        row["_row_number"] = excel_row_num
        rows.append(row)

    return rows
